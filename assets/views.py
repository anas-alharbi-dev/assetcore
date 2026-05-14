from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from assets.models import *
from .serializers import AssetSerializer, EmployeeSerializer, AssignmentSerializer
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response  
from django.http import HttpResponse
import pandas as pd
import openpyxl
from django.http import JsonResponse


class AssetViewSet(viewsets.ModelViewSet):
    queryset = Asset.objects.all()
    serializer_class = AssetSerializer
    permission_classes = [IsAuthenticated]


    def get_queryset(self):
        user = self.request.user

        if user.is_staff:
            return Asset.objects.all()

        try:
            employee = user.employee
            return Asset.objects.filter(assignments__employee=employee).distinct()
        except Employee.DoesNotExist:
            return Asset.objects.none()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user

        # الأدمن يشوف كل الموظفين
        if user.is_staff:
            return Employee.objects.all()
        return Employee.objects.filter(user=user)


class AssignmentViewSet(viewsets.ModelViewSet):
    queryset = Assignment.objects.all()
    serializer_class = AssignmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user

        # الأدمن يشوف كل عمليات الإسناد
        if user.is_staff:
            return Assignment.objects.all()

        # المستخدم العادي يشوف فقط الإسنادات الخاصة به
        try:
            employee = user.employee
            return Assignment.objects.filter(employee=employee)
        except Employee.DoesNotExist:
            return Assignment.objects.none()



    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['available_assets'] = Asset.objects.filter(is_assigned=False)
        return context
    

    


@api_view(["GET"])
def export_assets_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Assets"

    sheet.append(["Name", "Type", "Serial Number", "Asset Tag", "Status"])

    assets = Asset.objects.all()

    for asset in assets:
        sheet.append([
            asset.name,
            asset.device_type,
            asset.serial_number,
            asset.asset_tag,
            "Assigned" if asset.is_assigned else "Available"
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = "attachment; filename=assets.xlsx"

    workbook.save(response)

    return response

@api_view(['GET'])
def reports_summary(request):
    total_assets = Asset.objects.count()
    assigned_assets = Asset.objects.filter(is_assigned=True).count()
    available_assets = Asset.objects.filter(is_assigned=False).count()
    total_employees = Employee.objects.count()

    return Response({
        "total_assets": total_assets,
        "assigned_assets": assigned_assets,
        "available_assets": available_assets,
        "total_employees": total_employees,
    })


@api_view(['GET'])
def export_assignments_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Assignments"

    sheet.append(["Asset", "Employee", "Assigned At", "Returned At"])

    assignments = Assignment.objects.select_related('asset', 'employee')

    for a in assignments:
        sheet.append([
            a.asset.name,
            a.employee.name,
            str(a.assigned_at),
            str(a.returned_at) if a.returned_at else ""
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response["Content-Disposition"] = "attachment; filename=assignments.xlsx"

    workbook.save(response)
    return response



@api_view(['GET'])
def export_employees_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Employees"

    sheet.append(["Name", "Employee ID"])

    employees = Employee.objects.all()

    for e in employees:
        sheet.append([e.name, e.employee_id])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response["Content-Disposition"] = "attachment; filename=employees.xlsx"

    workbook.save(response)
    return response





@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_assets_excel(request):
    file = request.FILES.get('file')

    if not file:
        return JsonResponse({'error': 'No file provided'}, status=400)

    try:
        df = pd.read_excel(file)

        for _, row in df.iterrows():
            Asset.objects.create(
    name=row.get('Name'),
    device_type=row.get('Type'),
    serial_number=row.get('Serial Num'),   # 👈 مهم
    asset_tag=row.get('Asset Tag'),
    status=row.get('Status', 'available')
)

        return JsonResponse({'message': 'Assets imported successfully ✅'})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_assets(request):
    file = request.FILES.get('file')

    if not file:
        return Response({"error": "No file uploaded"}, status=400)

    try:
        df = pd.read_excel(file)

        created_count = 0
        errors = []

        for index, row in df.iterrows():
            row_number = index + 2  # عشان الإكسل يبدأ من 1 + header

            name = row.get('name')
            device_type = row.get('device_type')

            if pd.isna(name):
                errors.append(f"Row {row_number}: name is missing")
                continue

            if pd.isna(device_type):
                errors.append(f"Row {row_number}: device_type is missing")
                continue

            Asset.objects.create(
                name=name,
                device_type=device_type,
                serial_number=row.get('serial_number'),
                asset_tag=row.get('asset_tag'),
                purchase_date=row.get('purchase_date') if pd.notna(row.get('purchase_date')) else None,
                is_assigned=False,
            )

            created_count += 1

        return Response({
            "created": created_count,
            "errors": errors
        }, status=201)

    except Exception as e:
        return Response({"error": str(e)}, status=400)